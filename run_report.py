"""
Run the Coaching Call Discussion Report and export to a single Excel sheet.
Layout matches the CareFirst Activity Report PDF format.

Usage:
    python3 run_report.py                    # all customers, all dates
    python3 run_report.py HP_SCCareFirst     # single customer
    python3 run_report.py HP_SCCareFirst 2025-04-01 2025-06-30  # customer + date range
"""
import sys
import os
sys.path.append(os.path.expanduser('~/Documents/dev/automation'))
from db_connect import get_connection
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# --- Parse args ---
customer_id = sys.argv[1] if len(sys.argv) > 1 else None
start_date = sys.argv[2] if len(sys.argv) > 2 else None
end_date = sys.argv[3] if len(sys.argv) > 3 else None

# --- Read SQL ---
sql_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'coaching_call_topics_goals.sql')
with open(sql_path, 'r') as f:
    full_sql = f.read()

# --- Inject filters if specified ---
if customer_id or start_date:
    filters = []
    if customer_id:
        filters.append(f"AND MC.CUSTOMERID = '{customer_id}'")
    if start_date:
        filters.append(f"AND TRUNC(MC.ENCOUNTERDATETIME)::DATE >= '{start_date}'")
    if end_date:
        filters.append(f"AND TRUNC(MC.ENCOUNTERDATETIME)::DATE <= '{end_date}'")
    # Insert after the OUTBOUND filter in STEP 1A
    inject_point = "AND UPPER(MC.DIRECTION) = 'OUTBOUND'"
    full_sql = full_sql.replace(inject_point, inject_point + '\n      ' + '\n      '.join(filters))

# --- Split and classify statements ---
raw_stmts = [s.strip() for s in full_sql.split(';') if s.strip()]
output_queries = []
setup_statements = []
for stmt in raw_stmts:
    first_keyword = None
    for line in stmt.split('\n'):
        stripped = line.strip()
        if stripped and not stripped.startswith('--'):
            first_keyword = stripped.split()[0].upper()
            break
    if first_keyword == 'SELECT':
        output_queries.append(stmt)
    elif first_keyword in ('DROP', 'CREATE'):
        setup_statements.append(stmt)

scope = customer_id or 'ALL CUSTOMERS'
print(f"Scope: {scope}")
if start_date:
    print(f"Date range: {start_date} to {end_date or 'present'}")
print(f"Setup: {len(setup_statements)} statements, Outputs: {len(output_queries)} queries")

# --- Execute ---
with get_connection() as conn:
    cursor = conn.cursor()

    print("\nRunning setup...")
    for i, stmt in enumerate(setup_statements):
        for line in stmt.split('\n'):
            if line.strip() and not line.strip().startswith('--'):
                print(f"  [{i+1}/{len(setup_statements)}] {line.strip()[:60]}...")
                break
        cursor.execute(stmt)

    print("\nRunning outputs...")
    names = ['Topics', 'Tier Breakdown', 'Tobacco', 'Goal Type & Domain',
             'Goal Domain & Status', 'Goal Number & Status', 'Goal Type & Status', 'Validation']
    dfs = {}
    for i, query in enumerate(output_queries):
        name = names[i] if i < len(names) else f'Output {i+1}'
        dfs[name] = pd.read_sql(query, conn)
        print(f"  {name}: {len(dfs[name])} rows")

# --- Build single-sheet Excel ---
suffix = f"_{customer_id}" if customer_id else ""
output_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), f'coaching_call_discussions{suffix}.xlsx')

wb = Workbook()
ws = wb.active
ws.title = 'Coaching Call Discussions'

# Styles
section_font = Font(bold=True, size=11, underline='single')
table_header_font = Font(bold=True, size=10)
table_header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def write_section(ws, start_row, start_col, title, df):
    """Write a titled table at the given position. Returns next available row."""
    ws.cell(row=start_row, column=start_col, value=title).font = section_font
    for c, col_name in enumerate(df.columns):
        cell = ws.cell(row=start_row + 1, column=start_col + c, value=col_name)
        cell.font = table_header_font
        cell.fill = table_header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    for r, row in df.iterrows():
        for c, val in enumerate(row):
            cell = ws.cell(row=start_row + 2 + r, column=start_col + c, value=val)
            cell.border = thin_border
            if isinstance(val, (int, float)):
                cell.alignment = Alignment(horizontal='right')
    return start_row + 2 + len(df) + 2

# Title
ws.cell(row=1, column=1, value='Coaching Call Discussions').font = Font(bold=True, size=14)
if customer_id:
    ws.cell(row=2, column=1, value=f'Customer: {customer_id}').font = Font(size=10)
if start_date:
    ws.cell(row=2, column=4, value=f'{start_date} to {end_date or "present"}').font = Font(size=10)

# Layout: Topics + Tobacco side by side, then goals below
row = 4
next_row = write_section(ws, row, 1, 'Topics', dfs['Topics'])
write_section(ws, row, 7, 'Tobacco', dfs['Tobacco'])

row = next_row
next_row = write_section(ws, row, 1, 'Goal Type & Domain', dfs['Goal Type & Domain'])

row = next_row
write_section(ws, row, 1, 'Goal Number & Status', dfs['Goal Number & Status'])
write_section(ws, row, 9, 'Goal Type & Status', dfs['Goal Type & Status'])
next_row = write_section(ws, row, 17, 'Goal Domains & Status', dfs['Goal Domain & Status'])

row = next_row
write_section(ws, row, 1, 'Topic Source Breakdown', dfs['Tier Breakdown'])

# Column widths
for col in range(1, 25):
    ws.column_dimensions[get_column_letter(col)].width = 14

wb.save(output_path)
print(f"\nDone! Saved to: {output_path}")
